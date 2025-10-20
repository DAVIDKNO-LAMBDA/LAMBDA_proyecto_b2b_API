from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from django.conf import settings
from django.utils import timezone
from django.core.files.base import ContentFile
import io
import os
import logging

logger = logging.getLogger(__name__)


class PDFGenerator:
    """
    Generador de PDFs para facturas y reportes usando ReportLab
    """
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """Configurar estilos personalizados"""
        
        # Estilo para títulos principales
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=20,
            alignment=1  # Centrado
        ))
        
        # Estilo para subtítulos
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12
        ))
        
        # Estilo para información de empresa
        self.styles.add(ParagraphStyle(
            name='CompanyInfo',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#7f8c8d'),
            spaceAfter=6
        ))
        
        # Estilo para totales
        self.styles.add(ParagraphStyle(
            name='TotalStyle',
            parent=self.styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#2c3e50'),
            fontName='Helvetica-Bold'
        ))
    
    def generar_factura_pdf(self, factura):
        """
        Genera PDF de factura (HU24)
        """
        try:
            # Crear buffer en memoria
            buffer = io.BytesIO()
            
            # Crear documento PDF
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Elementos del PDF
            elements = []
            
            # Header con logo y datos de Lambda
            elements.extend(self._crear_header_factura())
            
            # Información del cliente
            elements.extend(self._crear_info_cliente_factura(factura))
            
            # Detalles de la factura
            elements.extend(self._crear_detalles_factura(factura))
            
            # Tabla de items
            elements.extend(self._crear_tabla_items_factura(factura))
            
            # Totales
            elements.extend(self._crear_totales_factura(factura))
            
            # Footer
            elements.extend(self._crear_footer_factura(factura))
            
            # Construir PDF
            doc.build(elements)
            
            # Obtener contenido
            pdf_content = buffer.getvalue()
            buffer.close()
            
            return pdf_content
            
        except Exception as e:
            logger.error(f"Error generando PDF de factura {factura.numero_factura}: {str(e)}")
            raise
    
    def _crear_header_factura(self):
        """Crear header con logo de Lambda"""
        elements = []
        
        # Título principal
        titulo = Paragraph("LAMBDA COMMERCE SOLUTIONS", self.styles['CustomTitle'])
        elements.append(titulo)
        
        # Información de la empresa
        info_empresa = [
            "NIT: 900.123.456-7",
            "Dirección: Calle 123 #45-67, Bogotá D.C.",
            "Teléfono: +57 (1) 234-5678",
            "Email: facturacion@lambda.com",
            "www.lambda.com"
        ]
        
        for info in info_empresa:
            p = Paragraph(info, self.styles['CompanyInfo'])
            elements.append(p)
        
        elements.append(Spacer(1, 20))
        
        return elements
    
    def _crear_info_cliente_factura(self, factura):
        """Crear información del cliente y factura"""
        elements = []
        
        # Datos en dos columnas
        data = [
            ['FACTURA DE VENTA', f'No. {factura.numero_factura}'],
            ['', ''],
            ['CLIENTE:', ''],
            [f'{factura.empresa_cliente.nombre}', f'Fecha: {factura.fecha_emision.strftime("%d/%m/%Y")}'],
            [f'NIT: {factura.empresa_cliente.nit}', f'Vencimiento: {factura.fecha_vencimiento.strftime("%d/%m/%Y") if factura.fecha_vencimiento else "N/A"}'],
            [f'{factura.empresa_cliente.direccion}', f'Pedido: {factura.pedido.numero_pedido}'],
            [f'Tel: {factura.empresa_cliente.telefono}', ''],
            [f'Email: {factura.empresa_cliente.correo_contacto}', ''],
        ]
        
        tabla = Table(data, colWidths=[3*inch, 2.5*inch])
        tabla.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 2), (0, 2), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(tabla)
        elements.append(Spacer(1, 20))
        
        return elements
    
    def _crear_detalles_factura(self, factura):
        """Crear detalles específicos de la factura"""
        elements = []
        
        # Condiciones de pago
        condiciones = f"Modalidad de Pago: {factura.pedido.get_modalidad_pago_display()}"
        if factura.pedido.modalidad_pago == 'diferido':
            condiciones += f" | Plazo: {factura.fecha_vencimiento.strftime('%d/%m/%Y')}"
        
        p = Paragraph(condiciones, self.styles['Normal'])
        elements.append(p)
        elements.append(Spacer(1, 15))
        
        return elements
    
    def _crear_tabla_items_factura(self, factura):
        """Crear tabla de items de la factura"""
        elements = []
        
        # Headers de la tabla
        headers = ['DESCRIPCIÓN', 'CANT.', 'PRECIO UNIT.', 'SUBTOTAL']
        data = [headers]
        
        # Items de la factura
        for item in factura.items.all():
            data.append([
                item.descripcion,
                f"{item.cantidad:,.0f}",
                f"${item.precio_unitario:,.0f}",
                f"${item.subtotal:,.0f}"
            ])
        
        # Crear tabla
        tabla = Table(data, colWidths=[3*inch, 0.8*inch, 1.2*inch, 1.2*inch])
        tabla.setStyle(TableStyle([
            # Headers
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Descripción alineada a la izquierda
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            
            # Alternar colores en filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        
        elements.append(tabla)
        elements.append(Spacer(1, 15))
        
        return elements
    
    def _crear_totales_factura(self, factura):
        """Crear sección de totales"""
        elements = []
        
        # Tabla de totales (alineada a la derecha)
        totales_data = [
            ['Subtotal:', f"${factura.subtotal:,.0f}"],
            ['Descuento:', f"${factura.descuento:,.0f}"],
            ['IVA (19%):', f"${factura.impuestos:,.0f}"],
            ['', ''],
            ['TOTAL:', f"${factura.total:,.0f}"]
        ]
        
        tabla_totales = Table(totales_data, colWidths=[1.5*inch, 1.2*inch])
        tabla_totales.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 3), 'Helvetica'),
            ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 3), 10),
            ('FONTSIZE', (0, 4), (-1, 4), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LINEBELOW', (0, 3), (-1, 3), 1, colors.black),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#ecf0f1')),
        ]))
        
        # Alinear tabla a la derecha
        tabla_container = Table([[tabla_totales]], colWidths=[6.2*inch])
        tabla_container.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ]))
        
        elements.append(tabla_container)
        elements.append(Spacer(1, 20))
        
        return elements
    
    def _crear_footer_factura(self, factura):
        """Crear footer de la factura"""
        elements = []
        
        # Observaciones si existen
        if factura.observaciones:
            elements.append(Paragraph("<b>Observaciones:</b>", self.styles['Normal']))
            elements.append(Paragraph(factura.observaciones, self.styles['Normal']))
            elements.append(Spacer(1, 10))
        
        # Términos y condiciones
        terminos = factura.terminos_condiciones or """
        TÉRMINOS Y CONDICIONES:
        1. Esta factura se considera aceptada si no se objeta dentro de los 8 días siguientes a su recepción.
        2. Los pagos tardíos causarán intereses de mora según la ley.
        3. Para cualquier aclaración, comunicarse al teléfono indicado.
        """
        
        elements.append(Paragraph("<b>Términos y Condiciones:</b>", self.styles['Normal']))
        elements.append(Paragraph(terminos, self.styles['CompanyInfo']))
        
        # Información de generación
        info_gen = f"Factura generada electrónicamente el {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(info_gen, self.styles['CompanyInfo']))
        
        return elements
    
    def generar_reporte_ventas_pdf(self, facturas, fecha_inicio, fecha_fin):
        """
        Genera PDF de reporte de ventas
        """
        try:
            # Crear buffer en memoria
            buffer = io.BytesIO()
            
            # Crear documento PDF
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Elementos del PDF
            elements = []
            
            # Título
            titulo = Paragraph(f"REPORTE DE VENTAS - {fecha_inicio} a {fecha_fin}", self.styles['CustomTitle'])
            elements.append(titulo)
            elements.append(Spacer(1, 20))
            
            # Tabla de facturas
            if facturas.exists():
                headers = ['Fecha', 'Factura', 'Cliente', 'Total']
                data = [headers]
                
                for factura in facturas:
                    data.append([
                        factura.fecha_emision.strftime('%d/%m/%Y'),
                        factura.numero_factura,
                        factura.empresa_cliente.nombre,
                        f"${factura.total:,.0f}"
                    ])
                
                tabla = Table(data, colWidths=[1.5*inch, 2*inch, 3*inch, 1.5*inch])
                tabla.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                elements.append(tabla)
            else:
                elements.append(Paragraph("No se encontraron datos para el período especificado.", self.styles['Normal']))
            
            # Construir PDF
            doc.build(elements)
            
            # Obtener contenido
            pdf_content = buffer.getvalue()
            buffer.close()
            
            return pdf_content
            
        except Exception as e:
            logger.error(f"Error generando PDF de reporte: {str(e)}")
            raise


class ExcelGenerator:
    """
    Generador de reportes en Excel usando openpyxl
    """
    
    def generar_reporte_ventas_excel(self, datos, fecha_inicio, fecha_fin):
        """Genera reporte de ventas en Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Ventas"
            
            # Título
            ws['A1'] = f"REPORTE DE VENTAS - {fecha_inicio} a {fecha_fin}"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ['Fecha', 'Factura', 'Cliente', 'Pedido', 'Subtotal', 'Descuento', 'IVA', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Datos
            for row, factura in enumerate(datos, 4):
                ws.cell(row=row, column=1, value=factura.fecha_emision)
                ws.cell(row=row, column=2, value=factura.numero_factura)
                ws.cell(row=row, column=3, value=factura.empresa_cliente.nombre)
                ws.cell(row=row, column=4, value=factura.pedido.numero_pedido)
                ws.cell(row=row, column=5, value=float(factura.subtotal))
                ws.cell(row=row, column=6, value=float(factura.descuento))
                ws.cell(row=row, column=7, value=float(factura.impuestos))
                ws.cell(row=row, column=8, value=float(factura.total))
            
            # Ajustar ancho de columnas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True
            
            # Guardar en buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando Excel: {str(e)}")
            raise


# Instancia global del generador
pdf_generator = PDFGenerator()
excel_generator = ExcelGenerator()